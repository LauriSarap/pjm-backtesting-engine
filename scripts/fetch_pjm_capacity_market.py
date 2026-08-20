#!/usr/bin/env python3
"""Fetch and normalize PJM RPM capacity-market workbooks. Requires: openpyxl."""

from __future__ import annotations

import argparse
import csv
import os
import re
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

# Data lives outside the package: $PJM_DATA_ROOT if set, else ./data
DATA_ROOT = Path(os.environ.get("PJM_DATA_ROOT", "data")).resolve()
DEFAULT_OUTPUT = DATA_ROOT / "raw" / "pjm" / "capacity_market"


FILES = {
    "rpm-auctions-resource-clearing-price-summary.xlsx": "https://www.pjm.com/-/media/DotCom/markets-ops/rpm/rpm-auction-info/rpm-auctions-resource-clearing-price-summary.xlsx",
    "rpm-commitment-by-fuel-type-by-dy-27-28.xlsx": "https://www.pjm.com/-/media/DotCom/markets-ops/rpm/rpm-auction-info/rpm-commitment-by-fuel-type-by-dy-27-28.xlsx",
    "rpm-auction-schedule.xlsx": "https://www.pjm.com/-/media/DotCom/markets-ops/rpm/rpm-auction-info/rpm-auction-schedule.xlsx",
    "pai-balancing-ratios-20191004.xlsx": "https://www.pjm.com/-/media/DotCom/markets-ops/rpm/rpm-auction-info/pai-balancing-ratios-20191004.xlsx",
    "2022-2023-pai-charge-rates-by-lda.xlsx": "https://www.pjm.com/-/media/DotCom/markets-ops/rpm/2022-2023-pai-charge-rates-by-lda.xlsx",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fetch and normalize PJM RPM capacity-market workbooks."
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--skip-download", action="store_true")
    return parser.parse_args()


def download_files(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for filename, url in FILES.items():
        output_path = output_dir / filename
        print(f"downloading {filename}", flush=True)
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "pjm-backtesting-engine"},
        )
        with urllib.request.urlopen(request, timeout=120) as response:
            output_path.write_bytes(response.read())


def clean_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def as_float(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"wrote {path} ({len(rows)} rows)", flush=True)


def normalize_resource_clearing_prices(output_dir: Path) -> None:
    path = output_dir / "rpm-auctions-resource-clearing-price-summary.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["RCP"]

    headers = [clean_cell(v) for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    lda_headers = headers[2:25]
    rows: list[dict[str, Any]] = []
    delivery_year = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        first = clean_cell(row[0])
        if isinstance(first, str) and re.fullmatch(r"DY\s*\d{2}/\d{2}", first):
            year = re.search(r"(\d{2})/(\d{2})", first)
            if not year:
                raise RuntimeError(f"Could not parse delivery year from {first}")
            delivery_year = f"20{year.group(1)}/20{year.group(2)}"
            continue
        if not delivery_year or not first:
            continue
        auction = str(first)
        product_type = clean_cell(row[1])
        for offset, lda in enumerate(lda_headers, start=2):
            raw_value = clean_cell(row[offset])
            if lda is None or raw_value in (None, ""):
                continue
            rows.append(
                {
                    "delivery_year": delivery_year,
                    "auction": auction,
                    "capacity_product_type": product_type,
                    "lda": lda,
                    "resource_clearing_price_mw_day": as_float(raw_value),
                    "raw_value": raw_value,
                }
            )

    write_csv(
        output_dir / "resource_clearing_prices_long.csv",
        rows,
        [
            "delivery_year",
            "auction",
            "capacity_product_type",
            "lda",
            "resource_clearing_price_mw_day",
            "raw_value",
        ],
    )


def normalize_commitments(output_dir: Path) -> None:
    path = output_dir / "rpm-commitment-by-fuel-type-by-dy-27-28.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        headers = [
            clean_cell(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        fuel_headers = headers[2:]
        delivery_year = None
        unit = "UCAP" if "UCAP" in ws.title.upper() else "ICAP"
        for row in ws.iter_rows(min_row=2, values_only=True):
            first = clean_cell(row[0])
            if first:
                delivery_year = str(first)
            data_type = clean_cell(row[1])
            if not delivery_year or not data_type:
                continue
            for offset, fuel_type in enumerate(fuel_headers, start=2):
                if fuel_type in (None, ""):
                    continue
                mw = as_float(row[offset])
                if mw is None:
                    continue
                rows.append(
                    {
                        "delivery_year": delivery_year,
                        "unit": unit,
                        "data_type": data_type,
                        "fuel_type": fuel_type,
                        "mw": mw,
                    }
                )

    write_csv(
        output_dir / "commitments_by_fuel_type_long.csv",
        rows,
        ["delivery_year", "unit", "data_type", "fuel_type", "mw"],
    )


def normalize_pai_balancing_ratios(output_dir: Path) -> None:
    path = output_dir / "pai-balancing-ratios-20191004.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Balancing Ratios"]
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[0] is None:
            continue
        rows.append(
            {
                "ref_id": clean_cell(row[0]),
                "pai_date": clean_cell(row[1]),
                "pai_area": clean_cell(row[2]),
                "pai_start_time_gmt": clean_cell(row[3]),
                "pai_stop_time_gmt": clean_cell(row[4]),
                "pai_start_time_ept": clean_cell(row[5]),
                "pai_stop_time_ept": clean_cell(row[6]),
                "preliminary_balancing_ratio": as_float(row[7]),
                "preliminary_br_posting_time_ept": clean_cell(row[8]),
                "final_balancing_ratio": as_float(row[9]),
            }
        )

    write_csv(
        output_dir / "pai_balancing_ratios.csv",
        rows,
        [
            "ref_id",
            "pai_date",
            "pai_area",
            "pai_start_time_gmt",
            "pai_stop_time_gmt",
            "pai_start_time_ept",
            "pai_stop_time_ept",
            "preliminary_balancing_ratio",
            "preliminary_br_posting_time_ept",
            "final_balancing_ratio",
        ],
    )


def normalize_pai_charge_rates(output_dir: Path) -> None:
    path = output_dir / "2022-2023-pai-charge-rates-by-lda.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        rows.append(
            {
                "delivery_year": clean_cell(row[0]),
                "lda": clean_cell(row[1]),
                "net_cone_icap": as_float(row[2]),
                "pai_charge_rate": as_float(row[3]),
            }
        )
    write_csv(
        output_dir / "pai_charge_rates_by_lda.csv",
        rows,
        ["delivery_year", "lda", "net_cone_icap", "pai_charge_rate"],
    )


def normalize_auction_schedule(output_dir: Path) -> None:
    path = output_dir / "rpm-auction-schedule.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        header_row = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and row[0] == "Delivery Year" and row[1] == "Auction":
                header_row = idx
                break
        if header_row is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row[0]:
                continue
            rows.append(
                {
                    "sheet": ws.title,
                    "delivery_year": clean_cell(row[0]),
                    "auction": clean_cell(row[1]),
                    "activity_type": clean_cell(row[2]),
                    "activity": clean_cell(row[3]),
                    "responsible_party": clean_cell(row[4]),
                    "date": clean_cell(row[5]),
                }
            )
    write_csv(
        output_dir / "auction_schedule.csv",
        rows,
        [
            "sheet",
            "delivery_year",
            "auction",
            "activity_type",
            "activity",
            "responsible_party",
            "date",
        ],
    )


def main() -> None:
    args = parse_args()
    if not args.skip_download:
        download_files(args.output_dir)
    normalize_resource_clearing_prices(args.output_dir)
    normalize_commitments(args.output_dir)
    normalize_pai_balancing_ratios(args.output_dir)
    normalize_pai_charge_rates(args.output_dir)
    normalize_auction_schedule(args.output_dir)


if __name__ == "__main__":
    main()
